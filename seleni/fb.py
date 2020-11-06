
import csv
import time
# from bs4 import BeautifulSoup
import openpyxl
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
# from twocaptcha import TwoCaptcha

# начало сохранения картинок надо доработать
# into_link = requests.get('https://static.xx.fbcdn.net/rsrc.php/v3/ya/r/mTNr7Jn2-Jk.png', timeout=20)
# # print(into_link.content)
# with open(f'img2.png', 'wb') as output:
#     output.write(into_link.content)
# wb = Workbook()
# # wb = openpyxl.load_workbook('img2.xlsx')
# wb.create_sheet('Sheet1')
# worksheet = wb['Sheet1']
# img = Image('img2.png')
# img.anchor = 'A2'
# worksheet['A1'] = 'assss'
# worksheet.add_image(img)
# wb.save('qqq.xlsx')

columns = ['post', 'profile_link', 'fio', 'text', 'date', 'answer']
with open('fb_parser_2.csv', 'w', newline='') as file:
    writer = csv.writer(file, delimiter='|')
    writer.writerow(columns)

post_link = "https://www.facebook.com/akipress/posts/3447581221984882"
comments_all = []
option = Options()

option.add_argument("--disable-infobars")
option.add_argument("start-maximized")
option.add_argument("--disable-extensions")

# Pass the argument 1 to allow and 2 to block
option.add_experimental_option("prefs", {
    "profile.default_content_setting_values.notifications": 1
})

driver = webdriver.Chrome(chrome_options=option)
# driver.set_script_timeout(1)
# driver.set_page_load_timeout(1)
# driver.implicitly_wait(1)
# открываем ссылку страницы входа
print('открываем ссылку страницы входа')
# driver.get("https://www.facebook.com/japarov.sadyr/posts/1684394035068852")
driver.get("https://www.facebook.com")

driver.implicitly_wait(10)
time.sleep(2)

WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'button[data-cookiebanner="accept_button"]'))).click()

driver.implicitly_wait(10)
time.sleep(2)

driver.find_element_by_name('email').send_keys('akyl.aydarbekov@gmail.com')
driver.find_element_by_name('pass').send_keys('sprite05')
driver.find_element_by_css_selector('button[type="submit"]').click()

driver.implicitly_wait(10)
time.sleep(2)

driver.get(post_link)

driver.implicitly_wait(10)
time.sleep(2)
comments_num = 0
# 88888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888888
post_type = ''
try:
    driver.find_element_by_xpath("//*[contains(text(), 'Свяжитесь с нами')]")
    post_type = 'NEWS'
except NoSuchElementException as e:
    try:
        driver.find_element_by_xpath("//*[contains(text(), 'Сообщение')]")
        post_type = 'NEWS'
    except NoSuchElementException as e:
        try:
            driver.find_element_by_xpath("//*[contains(text(), 'Смотреть видео')]")
            post_type = 'NEWS'
        except NoSuchElementException as e:
            try:
                driver.find_element_by_xpath("//*[contains(text(), 'Использовать приложение')]")
                post_type = 'NEWS'
            except:
                print('It is not NEWS')
                print('Trying find GROUP')
                try:
                    driver.find_element_by_xpath("//*[contains(text(), 'Вступить в группу')]")
                    post_type = 'GROUP'
                except NoSuchElementException as e:
                    print('It is not GROUP')
                    print('Trying find NEWS')
                    try:
                        driver.find_element_by_xpath("//*[contains(text(), 'Самые актуальные')]")
                        post_type = 'SIMPLE'
                    except NoSuchElementException as e:
                        print('It is not SIMPLE')
                        print('WHATAFUCKA')


    # mas, d = (e.msg).split('\n')
    # if mas == 'no such element: Unable to locate element: {"method":"xpath","selector":"//*[contains(text(), \'Самые актуальные\')]"}':
    #     print('its group or news')


if post_type == 'SIMPLE':
    driver.find_element_by_xpath("//*[contains(text(), 'Самые актуальные')]").click()
    time.sleep(3)
    driver.find_element_by_xpath("//*[contains(text(), 'От новых к старым')]").click()
    time.sleep(3)

    while True:
        # driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        more_comments = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Показать')]")))

        time.sleep(3)
        try:
            sp = driver.find_element_by_xpath("//*[contains(text(), '4 из ')]")
        except NoSuchElementException:
            print('all comments')
            break
        print(sp.text)
        sp = (sp.text).split(' из ')
        print(sp)
        print(int(sp[0]))
        print(comments_num)
        print(int(sp[0].replace(' ', '')) != comments_num)
        if int(sp[0].replace(' ', '')) != comments_num:
            comments_num = int(sp[0])
            if more_comments:
                more_comments.click()
                time.sleep(10)
            else:
                break
        else:
            time.sleep(10)
            if more_comments:
                more_comments.click()
                time.sleep(10)
        # if more_comments:
        #     more_comments.click()
        #     time.sleep(10)

    print('thats all comments')

    comments = driver.find_elements_by_xpath("//div[@class='cwj9ozl2 tvmbv18p']/ul/li")
    # tw6a2znq sj5x9vvc d1544ag0 cxgpxx05
    for comment in comments:
        comment_list = []
        profile_link = comment.find_element_by_css_selector('a[class="oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl gmql0nx0 gpro0wi8"]')
        print(profile_link.get_attribute('href'))
        comment_list.append(post_link)
        comment_list.append(profile_link.get_attribute('href'))
        print(profile_link.text)
        comment_list.append(profile_link.text)

        try:
            # more = WebDriverWait(comment, 0).until(
                # EC.presence_of_element_located((By.XPATH, ".//*[contains(text(), 'Ещё')]")))
            more = comment.find_element_by_xpath(".//*[contains(text(), 'Ещё')]")
            driver.set_script_timeout(1)
            driver.implicitly_wait(1)
            # more = comment.find_element_by_xpath(".//*[contains(text(), 'Ещё')]")
            print('founded more')
            more.click()
            time.sleep(3)
        except:
            print('2')
            print('full comment')
        try:
            text = WebDriverWait(comment, 0.1).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div[class='ecm0bbzt e5nlhep0 a8c37x1j']")))
            # text = comment.find_element_by_css_selector('div[class="ecm0bbzt e5nlhep0 a8c37x1j"]')
            print(text.text)
            comment_list.append(text.text)
        except:
            text = 'no text'
            comment_list.append(text)

        time_com = comment.find_element_by_css_selector('a[class="oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl m9osqain gpro0wi8 knj5qynh"]')
        hover = ActionChains(driver).move_to_element(time_com)
        hover.perform()
        time.sleep(1)
        time_full = driver.find_element_by_css_selector('span[class="d2edcug0 hpfvmrgz qv66sw1b c1et5uql gk29lw5a a8c37x1j keod5gw0 nxhoafnm aigsh9s9 tia6h79c fe6kdd0r mau55g9w c8b282yb iv3no6db e9vueds3 j5wam9gi knj5qynh oo9gr5id hzawbc8m"]')
        print(time_full.text)
        comment_list.append(time_full.text)
        comment_list.append('---')
        # comments_all.append(comment_list)
        # print(comments_all)
        with open('fb_parser_2.csv', 'a+', newline='') as file:
            writer = csv.writer(file, delimiter='|')
            writer.writerow(comment_list)
        comment_list.clear()
        #  Ищем ответы
        try:
            answer_btn = WebDriverWait(comment, 0).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div[class="oajrlxb2 bp9cbjyn g5ia77u1 mtkw9kbi tlpljxtp qensuy8j ppp5ayq2 goun2846 ccm00jje s44p3ltw mk2mc5f4 rt8b4zig n8ej3o3l agehan2d sk4xxmp2 rq0escxv nhd2j8a9 pq6dq46d mg4g778l btwxx1t3 g5gj957u p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x tgvbjcpo hpfvmrgz jb3vyjys p8fzw8mz qt6c0cv9 a8nywdso l9j0dhe7 i1ao9s8h esuyzwwr f1sip0of du4w35lb lzcic4wl abiwlrkh gpro0wi8 m9osqain buofh1pr"]')))
            # answer_btn = comment.find_element_by_css_selector('div[class="oajrlxb2 bp9cbjyn g5ia77u1 mtkw9kbi tlpljxtp qensuy8j ppp5ayq2 goun2846 ccm00jje s44p3ltw mk2mc5f4 rt8b4zig n8ej3o3l agehan2d sk4xxmp2 rq0escxv nhd2j8a9 pq6dq46d mg4g778l btwxx1t3 g5gj957u p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x tgvbjcpo hpfvmrgz jb3vyjys p8fzw8mz qt6c0cv9 a8nywdso l9j0dhe7 i1ao9s8h esuyzwwr f1sip0of du4w35lb lzcic4wl abiwlrkh gpro0wi8 m9osqain buofh1pr"]')
            driver.set_script_timeout(1)
            driver.implicitly_wait(1)
            print('found answers')
            answer_btn.click()
            time.sleep(3)
            answers = comment.find_elements_by_xpath(".//div/div/ul/li/div/div")
            for ans in answers:
                ans_profile = ans.find_element_by_css_selector('a[class="oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl gmql0nx0 gpro0wi8"]')
                print(ans_profile.get_attribute('href'))
                comment_list = []
                comment_list.append(post_link)
                comment_list.append(ans_profile.get_attribute('href'))
                print(ans_profile.text)
                comment_list.append(ans_profile.text)
                try:

                    ans_more = WebDriverWait(ans, 0.1).until(
                        EC.presence_of_element_located((By.XPATH, ".//*[contains(text(), 'Ещё')]")))
                    driver.set_script_timeout(1)
                    driver.implicitly_wait(1)
                    # ans_more = ans.find_element_by_xpath(".//*[contains(text(), 'Ещё')]")
                    print('founded more')
                    ans_more.click()
                    time.sleep(3)
                except:
                    print('full answer')
                try:
                    ans_text = WebDriverWait(ans, 0.1).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div[class='ecm0bbzt e5nlhep0 a8c37x1j']")))
                    # text = comment.find_element_by_css_selector('div[class="ecm0bbzt e5nlhep0 a8c37x1j"]')
                    print(ans_text.text)
                    comment_list.append(ans_text.text)
                except:
                    text = 'no text'
                    comment_list.append(text)
                # ans_text = ans.find_element_by_css_selector('div[class="ecm0bbzt e5nlhep0 a8c37x1j"]')

                ans_time_com = ans.find_element_by_css_selector(
                    'a[class="oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl m9osqain gpro0wi8 knj5qynh"]')
                hover = ActionChains(driver).move_to_element(ans_time_com)
                hover.perform()
                print('hovered')
                time.sleep(2)
                ans_time_full = driver.find_element_by_css_selector(
                    'span[class="d2edcug0 hpfvmrgz qv66sw1b c1et5uql gk29lw5a a8c37x1j keod5gw0 nxhoafnm aigsh9s9 tia6h79c fe6kdd0r mau55g9w c8b282yb iv3no6db e9vueds3 j5wam9gi knj5qynh oo9gr5id hzawbc8m"]')
                print(ans_time_full.text)
                comment_list.append(ans_time_full.text)
                comment_list.append('answer')
                # comments_all.append(comment_list)
                # print(comments_all)
                with open('fb_parser_2.csv', 'a+', newline='') as file:
                    writer = csv.writer(file, delimiter='|')
                    writer.writerow(comment_list)
                comment_list.clear()
        except Exception as e:
            print(e)
            print('no answers')

elif post_type == 'GROUP':
    print('its group')
    while True:
        try:
            driver.find_element_by_xpath("//*[contains(text(), 'Показать предыдущие комментарии')]").click()
            time.sleep(5)
        except NoSuchElementException as e:
            print('thats all comments')
            break

    comments = driver.find_elements_by_xpath("//div[@class='cwj9ozl2 tvmbv18p']/ul/li")
    # # tw6a2znq sj5x9vvc d1544ag0 cxgpxx05
    for comment in comments:
        try:
            asd = (comment.text).split('\n')
        except AttributeError:
            break
        comment_list = []
        profile_link = comment.find_element_by_css_selector(
            'a[class="oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl gmql0nx0 gpro0wi8"]')
        print(profile_link.get_attribute('href'))
        comment_list.append(post_link)
        comment_list.append(profile_link.get_attribute('href'))
        print(asd[0])
        comment_list.append(asd[0])
        if 'Ещё' in comment.text:
            try:
                more = comment.find_element_by_xpath(".//*[contains(text(), 'Ещё')]")
                driver.set_script_timeout(1)
                driver.implicitly_wait(1)
                print('founded more')
                more.click()
                time.sleep(3)
            except:
                print('full comment')
        try:
            text = WebDriverWait(comment, 0.1).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div[class='ecm0bbzt e5nlhep0 a8c37x1j']")))
            text_a = text.text
            try:
                icons = comment.find_elements_by_css_selector("img")
                for icon in icons:
                    icon_src = icon.get_attribute('src')
                    if 'https://' in icon_src:
                        text_a += ' ' + icon_src
            except:
                pass
            try:
                gifs = comment.find_elements_by_css_selector("a[aria-label='Нажмите, чтобы посмотреть вложение']")
                for gif in gifs:
                    print(gif.get_attribute('href'))
                    gif_href = gif.get_attribute('href')
                    text_a += ' ' + gif_href
            except:
                pass

            print(text_a)

            comment_list.append(text_a)
        except:
            text = ''
            gifs = comment.find_elements_by_css_selector("a[aria-label='Нажмите, чтобы посмотреть вложение']")
            for gif in gifs:
                print(gif.get_attribute('href'))
                gif_href = gif.get_attribute('href')
                text += ' ' + gif_href
            comment_list.append(text)

        time_com = comment.find_element_by_css_selector(
            'a[class="oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl m9osqain gpro0wi8 knj5qynh"]')
        hover = ActionChains(driver).move_to_element(time_com)
        hover.perform()
        time.sleep(1)
        time_full = driver.find_element_by_css_selector(
            'span[class="d2edcug0 hpfvmrgz qv66sw1b c1et5uql gk29lw5a a8c37x1j keod5gw0 nxhoafnm aigsh9s9 tia6h79c fe6kdd0r mau55g9w c8b282yb iv3no6db e9vueds3 j5wam9gi knj5qynh oo9gr5id hzawbc8m"]')
        print(time_full.text)
        comment_list.append(time_full.text)
        comment_list.append('---')
        # comments_all.append(comment_list)
        # print(comments_all)
        with open('fb_parser_2.csv', 'a+', newline='') as file:
            writer = csv.writer(file, delimiter='|')
            writer.writerow(comment_list)
        comment_list.clear()
        #  Ищем ответы
        try:
            # answer_btn = WebDriverWait(comment, 0).until(
            #     EC.presence_of_element_located((By.CSS_SELECTOR,
            #                                     'div[class="oajrlxb2 bp9cbjyn g5ia77u1 mtkw9kbi tlpljxtp qensuy8j ppp5ayq2 goun2846 ccm00jje s44p3ltw mk2mc5f4 rt8b4zig n8ej3o3l agehan2d sk4xxmp2 rq0escxv nhd2j8a9 pq6dq46d mg4g778l btwxx1t3 g5gj957u p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x tgvbjcpo hpfvmrgz jb3vyjys p8fzw8mz qt6c0cv9 a8nywdso l9j0dhe7 i1ao9s8h esuyzwwr f1sip0of du4w35lb lzcic4wl abiwlrkh gpro0wi8 m9osqain buofh1pr"]')))
            # answer_btn = comment.find_element_by_css_selector('div[class="oajrlxb2 bp9cbjyn g5ia77u1 mtkw9kbi tlpljxtp qensuy8j ppp5ayq2 goun2846 ccm00jje s44p3ltw mk2mc5f4 rt8b4zig n8ej3o3l agehan2d sk4xxmp2 rq0escxv nhd2j8a9 pq6dq46d mg4g778l btwxx1t3 g5gj957u p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x tgvbjcpo hpfvmrgz jb3vyjys p8fzw8mz qt6c0cv9 a8nywdso l9j0dhe7 i1ao9s8h esuyzwwr f1sip0of du4w35lb lzcic4wl abiwlrkh gpro0wi8 m9osqain buofh1pr"]')
            # driver.set_script_timeout(1)
            # driver.implicitly_wait(1)
            # print('found answers')
            # answer_btn.click()
            # time.sleep(3)
            answers = comment.find_elements_by_xpath(".//div/div/ul/li/div/div")
            for ans in answers:
                asf = (ans.text).split('\n')

                ans_profile = ans.find_element_by_css_selector(
                    'a[class="oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl gmql0nx0 gpro0wi8"]')
                print(ans_profile.get_attribute('href'))
                print(asf[0])
                comment_list = []
                comment_list.append(post_link)
                comment_list.append(ans_profile.get_attribute('href'))
                comment_list.append(asf[0])
                if 'Ещё' in ans.text:
                    try:

                        ans_more = WebDriverWait(ans, 0.1).until(
                            EC.presence_of_element_located((By.XPATH, ".//*[contains(text(), 'Ещё')]")))
                        driver.set_script_timeout(1)
                        driver.implicitly_wait(1)
                        # ans_more = ans.find_element_by_xpath(".//*[contains(text(), 'Ещё')]")
                        print('founded more')
                        ans_more.click()
                        time.sleep(3)
                    except:
                        print('full answer')
                try:
                    ans_text = WebDriverWait(ans, 0.1).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div[class='ecm0bbzt e5nlhep0 a8c37x1j']")))
                    # text = comment.find_element_by_css_selector('div[class="ecm0bbzt e5nlhep0 a8c37x1j"]')
                    text_b = ans_text.text
                    try:
                        icons = ans.find_elements_by_css_selector("img")
                        for icon in icons:
                            icon_src = icon.get_attribute('src')
                            if 'https://' in icon_src:
                                text_b += ' ' + icon_src
                    except:
                        pass
                    try:
                        gifs = comment.find_elements_by_css_selector(
                            "a[aria-label='Нажмите, чтобы посмотреть вложение']")
                        for gif in gifs:
                            print(gif.get_attribute('href'))
                            gif_href = gif.get_attribute('href')
                            text_b += ' ' + gif_href
                    except:
                        pass

                    print(text_b)
                    comment_list.append(text_b)
                except:
                    text = ''
                    gifs = comment.find_elements_by_css_selector("a[aria-label='Нажмите, чтобы посмотреть вложение']")
                    for gif in gifs:
                        print(gif.get_attribute('href'))
                        gif_href = gif.get_attribute('href')
                        text += ' ' + gif_href
                    comment_list.append(text)
                # ans_text = ans.find_element_by_css_selector('div[class="ecm0bbzt e5nlhep0 a8c37x1j"]')

                ans_time_com = ans.find_element_by_css_selector(
                    'a[class="oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl m9osqain gpro0wi8 knj5qynh"]')
                hover = ActionChains(driver).move_to_element(ans_time_com)
                hover.perform()
                time.sleep(1)
                ans_time_full = driver.find_element_by_css_selector(
                    'span[class="d2edcug0 hpfvmrgz qv66sw1b c1et5uql gk29lw5a a8c37x1j keod5gw0 nxhoafnm aigsh9s9 tia6h79c fe6kdd0r mau55g9w c8b282yb iv3no6db e9vueds3 j5wam9gi knj5qynh oo9gr5id hzawbc8m"]')
                print(ans_time_full.text)
                comment_list.append(ans_time_full.text)
                comment_list.append('answer')
                # comments_all.append(comment_list)
                # print(comments_all)
                with open('fb_parser_2.csv', 'a+', newline='') as file:
                    writer = csv.writer(file, delimiter='|')
                    writer.writerow(comment_list)
                comment_list.clear()
        except Exception as e:
            print(e)
            print('no answers')

if post_type == 'NEWS':
    first_post = driver.find_element_by_css_selector(
        'div[class="d2edcug0 oh7imozk tr9rh885 abvwweq7 ejjq64ki"]')
    try:
        first_post.find_element_by_xpath(".//*[contains(text(), 'Самые актуальные')]").click()
        time.sleep(3)
        driver.find_element_by_xpath(".//*[contains(text(), 'От новых к старым')]").click()
        time.sleep(3)
    except:
        pass

    while True:
        # driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        try:
            more_comments = WebDriverWait(first_post, 15).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Показать')]")))
        except:
            pass
        time.sleep(3)
        try:
            sp = first_post.find_element_by_xpath(".//*[contains(text(), '2 из ')]")
        except NoSuchElementException:
            print('all comments')
            break
        print(sp.text)
        sp = (sp.text).split(' из ')
        print(sp)
        print(int(sp[0]))
        print(comments_num)
        print(int(sp[0].replace(' ', '')) != comments_num)
        if int(sp[0].replace(' ', '')) != comments_num:
            comments_num = int(sp[0])
            if more_comments:
                more_comments.click()
                time.sleep(10)
            else:
                break
        else:
            time.sleep(10)
            if more_comments:
                more_comments.click()
                time.sleep(10)
        # if more_comments:
        #     more_comments.click()
        #     time.sleep(10)

    print('thats all comments')

    comments = first_post.find_elements_by_xpath(".//div[@class='cwj9ozl2 tvmbv18p']/ul/li")
    for comment in comments:
        comment_list = []
        profile_link = comment.find_element_by_css_selector('a[class="oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl gmql0nx0 gpro0wi8"]')
        print(profile_link.get_attribute('href'))
        comment_list.append(post_link)
        comment_list.append(profile_link.get_attribute('href'))
        print(profile_link.text)
        comment_list.append(profile_link.text)

        try:
            # more = WebDriverWait(comment, 0).until(
                # EC.presence_of_element_located((By.XPATH, ".//*[contains(text(), 'Ещё')]")))
            more = comment.find_element_by_xpath(".//*[contains(text(), 'Ещё')]")
            driver.set_script_timeout(1)
            driver.implicitly_wait(1)
            # more = comment.find_element_by_xpath(".//*[contains(text(), 'Ещё')]")
            print('founded more')
            more.click()
            time.sleep(2)
        except:
            print('full comment')
        try:
            text = WebDriverWait(comment, 0.1).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div[class='ecm0bbzt e5nlhep0 a8c37x1j']")))
            text_a = text.text
            try:
                icons = comment.find_elements_by_css_selector("img")
                for icon in icons:
                    icon_src = icon.get_attribute('src')
                    if 'https://' in icon_src:
                        text_a += ' ' + icon_src
            except:
                pass
            try:
                gifs = comment.find_elements_by_css_selector("a[aria-label='Нажмите, чтобы посмотреть вложение']")
                for gif in gifs:
                    print(gif.get_attribute('href'))
                    gif_href = gif.get_attribute('href')
                    text_a += ' ' + gif_href
            except:
                pass
            print(text_a)
            comment_list.append(text_a)
        except:
            text = ''
            gifs = comment.find_elements_by_css_selector("a[aria-label='Нажмите, чтобы посмотреть вложение']")
            for gif in gifs:
                print(gif.get_attribute('href'))
                gif_href = gif.get_attribute('href')
                text += ' ' + gif_href
            comment_list.append(text)

        time_com = comment.find_element_by_css_selector('a[class="oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl m9osqain gpro0wi8 knj5qynh"]')
        hover = ActionChains(driver).move_to_element(time_com)
        hover.perform()
        time.sleep(1)
        time_full = driver.find_element_by_css_selector('span[class="d2edcug0 hpfvmrgz qv66sw1b c1et5uql gk29lw5a a8c37x1j keod5gw0 nxhoafnm aigsh9s9 tia6h79c fe6kdd0r mau55g9w c8b282yb iv3no6db e9vueds3 j5wam9gi knj5qynh oo9gr5id hzawbc8m"]')
        print(time_full.text)
        comment_list.append(time_full.text)
        comment_list.append('---')
        # comments_all.append(comment_list)
        # print(comments_all)
        with open('fb_parser_2.csv', 'a+', newline='') as file:
            writer = csv.writer(file, delimiter='|')
            writer.writerow(comment_list)
        comment_list.clear()
        #  Ищем ответы
        try:
            answer_btn = WebDriverWait(comment, 0).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div[class="oajrlxb2 bp9cbjyn g5ia77u1 mtkw9kbi tlpljxtp qensuy8j ppp5ayq2 goun2846 ccm00jje s44p3ltw mk2mc5f4 rt8b4zig n8ej3o3l agehan2d sk4xxmp2 rq0escxv nhd2j8a9 pq6dq46d mg4g778l btwxx1t3 g5gj957u p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x tgvbjcpo hpfvmrgz jb3vyjys p8fzw8mz qt6c0cv9 a8nywdso l9j0dhe7 i1ao9s8h esuyzwwr f1sip0of du4w35lb lzcic4wl abiwlrkh gpro0wi8 m9osqain buofh1pr"]')))
            # answer_btn = comment.find_element_by_css_selector('div[class="oajrlxb2 bp9cbjyn g5ia77u1 mtkw9kbi tlpljxtp qensuy8j ppp5ayq2 goun2846 ccm00jje s44p3ltw mk2mc5f4 rt8b4zig n8ej3o3l agehan2d sk4xxmp2 rq0escxv nhd2j8a9 pq6dq46d mg4g778l btwxx1t3 g5gj957u p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x tgvbjcpo hpfvmrgz jb3vyjys p8fzw8mz qt6c0cv9 a8nywdso l9j0dhe7 i1ao9s8h esuyzwwr f1sip0of du4w35lb lzcic4wl abiwlrkh gpro0wi8 m9osqain buofh1pr"]')
            driver.set_script_timeout(1)
            driver.implicitly_wait(1)
            print('found answers')
            answer_btn.click()
            time.sleep(3)
            answers = comment.find_elements_by_xpath(".//div/div/ul/li/div/div")
            for ans in answers:
                ans_profile = ans.find_element_by_css_selector('a[class="oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl gmql0nx0 gpro0wi8"]')
                print(ans_profile.get_attribute('href'))
                comment_list = []
                comment_list.append(post_link)
                comment_list.append(ans_profile.get_attribute('href'))
                print(ans_profile.text)
                comment_list.append(ans_profile.text)
                try:
                    ans_more = WebDriverWait(ans, 0.1).until(
                        EC.presence_of_element_located((By.XPATH, ".//*[contains(text(), 'Ещё')]")))
                    driver.set_script_timeout(1)
                    driver.implicitly_wait(1)
                    # ans_more = ans.find_element_by_xpath(".//*[contains(text(), 'Ещё')]")
                    print('founded more')
                    ans_more.click()
                    time.sleep(3)
                except:
                    print('full answer')
                try:
                    ans_text = WebDriverWait(ans, 0.1).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div[class='ecm0bbzt e5nlhep0 a8c37x1j']")))
                    # text = comment.find_element_by_css_selector('div[class="ecm0bbzt e5nlhep0 a8c37x1j"]')
                    text_b = ans_text.text
                    try:
                        icons = ans.find_elements_by_css_selector("img")
                        for icon in icons:
                            icon_src = icon.get_attribute('src')
                            if 'https://' in icon_src:
                                text_b += ' ' + icon_src
                    except:
                        pass
                    try:
                        gifs = comment.find_elements_by_css_selector(
                            "a[aria-label='Нажмите, чтобы посмотреть вложение']")
                        for gif in gifs:
                            print(gif.get_attribute('href'))
                            gif_href = gif.get_attribute('href')
                            text_b += ' ' + gif_href
                    except:
                        pass

                    print(text_b)
                    comment_list.append(text_b)
                except:
                    text = ''
                    gifs = comment.find_elements_by_css_selector("a[aria-label='Нажмите, чтобы посмотреть вложение']")
                    for gif in gifs:
                        print(gif.get_attribute('href'))
                        gif_href = gif.get_attribute('href')
                        text += ' ' + gif_href
                    comment_list.append(text)
                # ans_text = ans.find_element_by_css_selector('div[class="ecm0bbzt e5nlhep0 a8c37x1j"]')

                ans_time_com = ans.find_element_by_css_selector(
                    'a[class="oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl m9osqain gpro0wi8 knj5qynh"]')
                hover = ActionChains(driver).move_to_element(ans_time_com)
                hover.perform()
                print('hovered')
                time.sleep(2)
                ans_time_full = driver.find_element_by_css_selector(
                    'span[class="d2edcug0 hpfvmrgz qv66sw1b c1et5uql gk29lw5a a8c37x1j keod5gw0 nxhoafnm aigsh9s9 tia6h79c fe6kdd0r mau55g9w c8b282yb iv3no6db e9vueds3 j5wam9gi knj5qynh oo9gr5id hzawbc8m"]')
                print(ans_time_full.text)
                comment_list.append(ans_time_full.text)
                comment_list.append('answer')
                # comments_all.append(comment_list)
                # print(comments_all)
                with open('fb_parser_2.csv', 'a+', newline='') as file:
                    writer = csv.writer(file, delimiter='|')
                    writer.writerow(comment_list)
                comment_list.clear()
        except Exception as e:
            print(e)
            print('no answers')


# with open('fb_parser_2.csv', 'a+', newline='') as file:
#     writer = csv.writer(file, delimiter='|')
#     for com in comments_all:
#         writer.writerow(com)

